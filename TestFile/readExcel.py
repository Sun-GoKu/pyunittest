import os
import getpathInfo
from xlrd import open_workbook
import openpyxl



path = getpathInfo.get_Path()

class readExcel():
    def get_xlsx(self,xls_name,sheet_name):
        cls = []
        xlsPath = os.path.join(path,'case',xls_name)
        # file = open_workbook(xlsPath)
        # 由于格式支持问题，这里使用openpyxl来加载读取Excel文件
        file = openpyxl.load_workbook(xlsPath)
        sheet = file.get_sheet_by_name(sheet_name)

        # 获取最大row/col
        m_rows = sheet.max_row
        m_cols = sheet.max_column


        # 注意openpyxl遍历是从1开始的
        for i in range(2,m_rows+1):
            rows_value = []         # 先定义一个临时列表,保存每行的信息
            for j in range(1,m_cols+1):
                res = sheet.cell(i,j).value
                rows_value.append(res)

            cls.append(rows_value)

        return cls


if __name__ == "__main__":
    print(readExcel().get_xlsx('userCase.xlsx','login'))
    print(readExcel().get_xlsx('userCase.xlsx','login')[0][1])
    print(readExcel().get_xlsx('userCase.xlsx','login')[1][2])
